import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

# ==================== 创建工作簿 ====================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "SauceDemo测试用例"

# ==================== 完整的标准表头 ====================
headers = ["用例编号", "用例标题", "项目/模块", "优先级", "前置条件", "测试步骤", "测试数据", "预期结果", "执行状态", "执行人"]
ws.append(headers)

# ==================== 针对SauceDemo平台的详细用例数据 ====================
# 每条用例都写清楚了步骤和数据，完全符合你之前的期望
test_data = [
    {
        "id": "TC_LOGIN_001",
        "title": "标准用户正确登录",
        "module": "登录模块",
        "priority": "P0",
        "precondition": "已注册的有效账号（standard_user）",
        "steps": "1. 打开登录页\n2. 输入用户名\n3. 输入密码\n4. 点击登录按钮",
        "data": "用户名：standard_user\n密码：secret_sauce",
        "expected": "登录成功，跳转至商品列表页，URL包含/inventory.html"
    },
    {
        "id": "TC_LOGIN_002",
        "title": "错误密码登录失败",
        "module": "登录模块",
        "priority": "P1",
        "precondition": "已注册的有效账号",
        "steps": "1. 打开登录页\n2. 输入用户名\n3. 输入错误密码\n4. 点击登录按钮",
        "data": "用户名：standard_user\n密码：wrong_password",
        "expected": "登录失败，提示 'Username and password do not match'"
    },
    {
        "id": "TC_LOGIN_003",
        "title": "用户名为空登录失败",
        "module": "登录模块",
        "priority": "P1",
        "precondition": "无",
        "steps": "1. 打开登录页\n2. 不输入用户名\n3. 输入密码\n4. 点击登录按钮",
        "data": "用户名：空\n密码：secret_sauce",
        "expected": "登录失败，提示 'Username is required'"
    },
    {
        "id": "TC_CART_001",
        "title": "添加单个商品到购物车",
        "module": "购物车模块",
        "priority": "P0",
        "precondition": "已登录成功，处于商品列表页",
        "steps": "1. 找到商品'Sauce Labs Backpack'\n2. 点击对应的'Add to cart'按钮",
        "data": "商品名称：Sauce Labs Backpack",
        "expected": "右上角购物车徽章数字变为 1"
    },
    {
        "id": "TC_CART_002",
        "title": "从购物车移除商品",
        "module": "购物车模块",
        "priority": "P1",
        "precondition": "购物车中已有1个商品（Backpack）",
        "steps": "1. 点击右上角购物车图标进入购物车页\n2. 点击'Remove'按钮移除Backpack",
        "data": "商品名称：Sauce Labs Backpack",
        "expected": "购物车列表为空，右上角徽章消失（数字为0）"
    },
    {
        "id": "TC_CHECKOUT_001",
        "title": "完整结算流程成功",
        "module": "结算模块",
        "priority": "P0",
        "precondition": "已登录，购物车中已有至少1个商品",
        "steps": "1. 点击购物车图标\n2. 点击'Checkout'\n3. 输入First Name, Last Name, Postal Code\n4. 点击'Continue'\n5. 点击'Finish'",
        "data": "姓名：张三\n邮编：12345",
        "expected": "页面显示 'Thank you for your order'"
    },
    {
        "id": "TC_CHECKOUT_002",
        "title": "结算时邮编为空校验",
        "module": "结算模块",
        "priority": "P1",
        "precondition": "已登录，购物车中已有至少1个商品",
        "steps": "1. 点击购物车图标\n2. 点击'Checkout'\n3. 只填写First Name和Last Name，不填Postal Code\n4. 点击'Continue'",
        "data": "姓名：张三\n邮编：空",
        "expected": "页面提示 'Postal Code is required'"
    }
]

# 将数据写入Excel行
for tc in test_data:
    ws.append([
        tc["id"],
        tc["title"],
        tc["module"],
        tc["priority"],
        tc["precondition"],
        tc["steps"],
        tc["data"],
        tc["expected"],
        "",   # 执行状态（待后续自动化回填）
        ""    # 执行人（待填写）
    ])

# ==================== 专业级美化 ====================
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                     top=Side(style='thin'), bottom=Side(style='thin'))

col_widths = [12, 28, 16, 10, 25, 40, 30, 45, 14, 14]
for i, width in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# 应用样式到表头
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# 应用样式到数据行
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = cell_alignment
        cell.border = thin_border

# ==================== 保存文件 ====================
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SauceDemo完整测试用例.xlsx")
wb.save(file_path)
print(f"✅ 完整规范的测试用例已生成！")
print(f"📁 文件位置：{file_path}")
print("📌 现在你有了这份带完整列头的Excel，它可以作为你的求职作品集，后续我们也能让脚本读取它并自动执行。")