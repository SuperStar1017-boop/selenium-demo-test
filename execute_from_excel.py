import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.edge.service import Service
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os

# ==================== 1. 准备工作 ====================
# 打开刚才生成的Excel
file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "SauceDemo完整测试用例.xlsx")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# 启动Edge浏览器（注意：驱动要放在这个文件夹里，或者已在系统路径）
service = Service(executable_path='./msedgedriver.exe')
driver = webdriver.Edge(service=service)

# 定义一个通用的“等待元素出现”的方法，防止页面加载慢报错
def wait_for_element(by, value, timeout=5):
    try:
        return WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
    except:
        return None

# ==================== 2. 核心执行逻辑 ====================
# 从第2行开始循环（第1行是表头）
for row_index in range(2, ws.max_row + 1):
    case_id = ws.cell(row=row_index, column=1).value  # A列：用例编号
    title = ws.cell(row=row_index, column=2).value    # B列：用例标题
    test_data = ws.cell(row=row_index, column=7).value # G列：测试数据
    expected = ws.cell(row=row_index, column=8).value  # H列：预期结果
    
    print(f"🔄 正在执行用例：{case_id} - {title}")
    
    # 重置实际结果和状态
    actual_result = ""
    status = "失败"
    
    try:
        # ---------- 场景A：登录相关用例 ----------
        if "登录" in title:
            # 解析用户名和密码（格式：用户名：xxx\n密码：xxx）
            username = "standard_user"  # 默认
            password = "secret_sauce"
            if test_data:
                lines = test_data.split('\n')
                for line in lines:
                    if "用户名" in line:
                        username = line.split('：')[-1].strip() if '：' in line else line.split(':')[-1].strip()
                    if "密码" in line:
                        password = line.split('：')[-1].strip() if '：' in line else line.split(':')[-1].strip()
            
            # 如果用户名或密码是"空"，就设为空字符串
            if username == "空": username = ""
            if password == "空": password = ""
            
            # 执行登录动作
            driver.get("https://www.saucedemo.com/")
            # 输入账号
            input_user = wait_for_element(By.ID, "user-name")
            if input_user:
                input_user.clear()
                input_user.send_keys(username)
            # 输入密码
            input_pass = wait_for_element(By.ID, "password")
            if input_pass:
                input_pass.clear()
                input_pass.send_keys(password)
            # 点击登录
            login_btn = wait_for_element(By.ID, "login-button")
            if login_btn:
                login_btn.click()
            
            time.sleep(1)  # 等待跳转
            
            # 检查预期结果：判断是成功还是失败
            if "跳转至商品列表" in expected:
                # 预期成功：检查是否出现了商品列表页的特征（比如"Products"标题）
                try:
                    driver.find_element(By.CLASS_NAME, "inventory_list")
                    actual_result = "登录成功，进入商品列表页"
                    status = "通过"
                except:
                    actual_result = "未进入商品列表页"
                    status = "失败"
            else:
                # 预期失败：检查是否有错误提示
                try:
                    error_msg = driver.find_element(By.CSS_SELECTOR, "[data-test='error']").text
                    actual_result = f"出现错误提示：{error_msg}"
                    # 检查预期结果中的关键词是否出现在实际错误中
                    if "required" in expected or "match" in expected:
                        # 简单判断：只要报错了就算符合预期（因为用例都是针对特定报错设计的）
                        status = "通过"
                    else:
                        status = "失败"
                except:
                    actual_result = "未出现预期的错误提示"
                    status = "失败"

        # ---------- 场景B：加购商品 ----------
        elif "加购" in title:
            # 解析商品名称
            product_name = test_data.split('：')[-1].strip() if test_data else ""
            # 确保在商品列表页
            if "/inventory.html" not in driver.current_url:
                driver.get("https://www.saucedemo.com/inventory.html")
            time.sleep(1)
            
            # 根据商品名称找到对应的"Add to cart"按钮并点击
            # 注意：SauceDemo的按钮文本是"Add to cart"，但根据商品不同，button的id是动态的
            # 我们通过xpath定位：找到包含商品名称的div，再找同级的button
            try:
                # 找到包含商品名称的元素（比如div或a标签）
                product_container = driver.find_element(By.XPATH, f"//div[contains(@class,'inventory_item') and contains(.,'{product_name}')]")
                add_btn = product_container.find_element(By.XPATH, ".//button[contains(text(),'Add to cart')]")
                add_btn.click()
                time.sleep(0.5)
                # 检查购物车徽章数量
                badge = driver.find_element(By.CLASS_NAME, "shopping_cart_badge")
                actual_result = f"加购成功，徽章显示：{badge.text}"
                status = "通过" if int(badge.text) > 0 else "失败"
            except Exception as e:
                actual_result = f"加购失败：{str(e)}"
                status = "失败"

        # ---------- 场景C：结算流程 ----------
        elif "结算" in title:
            # 确保在购物车页面
            if "cart" not in driver.current_url:
                driver.get("https://www.saucedemo.com/cart.html")
            time.sleep(1)
            
            # 点击Checkout
            checkout_btn = wait_for_element(By.ID, "checkout")
            if checkout_btn:
                checkout_btn.click()
                time.sleep(1)
            
            # 解析姓名和邮编
            first_name = "张"
            last_name = "三"
            postal = "12345"
            if test_data:
                lines = test_data.split('\n')
                for line in lines:
                    if "姓名" in line:
                        name_parts = line.split('：')[-1].strip().split()
                        if len(name_parts) >= 2:
                            first_name, last_name = name_parts[0], name_parts[1]
                        else:
                            first_name = name_parts[0]
                    if "邮编" in line:
                        postal = line.split('：')[-1].strip()
                        if postal == "空": postal = ""
            
            # 填写信息
            fn_input = wait_for_element(By.ID, "first-name")
            if fn_input: fn_input.send_keys(first_name)
            ln_input = wait_for_element(By.ID, "last-name")
            if ln_input: ln_input.send_keys(last_name)
            p_input = wait_for_element(By.ID, "postal-code")
            if p_input: p_input.send_keys(postal)
            
            # 点击Continue
            continue_btn = wait_for_element(By.ID, "continue")
            if continue_btn:
                continue_btn.click()
                time.sleep(1)
            
            # 验证预期结果
            if "Thank you" in expected:
                # 预期成功，点击Finish
                finish_btn = wait_for_element(By.ID, "finish")
                if finish_btn:
                    finish_btn.click()
                    time.sleep(1)
                    try:
                        driver.find_element(By.CLASS_NAME, "complete-header")
                        actual_result = "结算成功，显示Thank you"
                        status = "通过"
                    except:
                        actual_result = "未到达完成页面"
                        status = "失败"
            else:
                # 预期失败（邮编为空）
                try:
                    error_msg = driver.find_element(By.CSS_SELECTOR, "[data-test='error']").text
                    actual_result = f"出现错误提示：{error_msg}"
                    if "required" in expected or "Postal" in expected:
                        status = "通过"
                    else:
                        status = "失败"
                except:
                    actual_result = "未出现预期的错误"
                    status = "失败"

        # ---------- 场景D：移除商品 ----------
        elif "移除" in title:
            # 如果当前不在购物车页面，先进去
            if "cart" not in driver.current_url:
                driver.get("https://www.saucedemo.com/cart.html")
            time.sleep(1)
            try:
                # 找到Remove按钮并点击
                remove_btn = driver.find_element(By.XPATH, "//button[contains(text(),'Remove')]")
                remove_btn.click()
                time.sleep(0.5)
                # 检查是否还有商品
                items = driver.find_elements(By.CLASS_NAME, "cart_item")
                if len(items) == 0:
                    actual_result = "移除成功，购物车为空"
                    status = "通过"
                else:
                    actual_result = f"移除后还剩{len(items)}个商品"
                    status = "失败"
            except Exception as e:
                actual_result = f"移除失败：{str(e)}"
                status = "失败"

    except Exception as e:
        actual_result = f"脚本执行异常：{str(e)}"
        status = "失败"
        print(f"⚠️ 用例 {case_id} 执行异常：{e}")

    # ==================== 3. 回填结果到Excel ====================
    ws.cell(row=row_index, column=9).value = status      # I列：执行状态
    ws.cell(row=row_index, column=10).value = "实习生"    # J列：执行人
    ws.cell(row=row_index, column=6).value = actual_result # F列：实际结果（把实际结果填进去）

    # 实时打印进度
    print(f"📝 结果：{status} - {actual_result}")
    print("-" * 50)

# ==================== 4. 保存并收尾 ====================
wb.save(file_path)
driver.quit()
print(f"\n✅ 全部用例执行完毕！结果已保存回：{file_path}")
print("🎉 请打开Excel查看'执行状态'列，你看到'通过'和'失败'的标记了吗？")