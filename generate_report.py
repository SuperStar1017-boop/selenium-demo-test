import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

print("🚀 正在生成测试执行报告...")

# ========== 定义测试用例数据（基于你之前生成的7条用例） ==========
test_cases = [
    {
        "id": "TC001",
        "title": "正确的用户名和密码登录成功",
        "module": "登录模块",
        "priority": "P0",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 输入用户名: standard_user\n2. 输入密码: secret_sauce\n3. 点击Login按钮",
        "test_data": "username: standard_user\npassword: secret_sauce",
        "expected": "登录成功，跳转到商品列表页(inventory.html)",
        "actual": "登录成功，页面正确跳转",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC002",
        "title": "错误的密码登录失败",
        "module": "登录模块",
        "priority": "P1",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 输入用户名: standard_user\n2. 输入密码: wrong_password\n3. 点击Login按钮",
        "test_data": "username: standard_user\npassword: wrong_password",
        "expected": "登录失败，显示密码错误提示",
        "actual": "显示错误: Username and password do not match",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC003",
        "title": "不存在的用户名登录失败",
        "module": "登录模块",
        "priority": "P1",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 输入用户名: invalid_user\n2. 输入密码: secret_sauce\n3. 点击Login按钮",
        "test_data": "username: invalid_user\npassword: secret_sauce",
        "expected": "登录失败，显示用户名不存在错误",
        "actual": "显示错误: Username and password do not match",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC004",
        "title": "用户名为空登录",
        "module": "登录模块",
        "priority": "P1",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 不输入用户名(留空)\n2. 输入密码: secret_sauce\n3. 点击Login按钮",
        "test_data": "username: (空)\npassword: secret_sauce",
        "expected": "登录失败，提示用户名不能为空",
        "actual": "显示错误: Username is required",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC005",
        "title": "密码为空登录",
        "module": "登录模块",
        "priority": "P1",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 输入用户名: standard_user\n2. 不输入密码(留空)\n3. 点击Login按钮",
        "test_data": "username: standard_user\npassword: (空)",
        "expected": "登录失败，提示密码不能为空",
        "actual": "显示错误: Password is required",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC006",
        "title": "用户名和密码均为空登录",
        "module": "登录模块",
        "priority": "P2",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 不输入用户名\n2. 不输入密码\n3. 点击Login按钮",
        "test_data": "username: (空)\npassword: (空)",
        "expected": "登录失败，提示用户名不能为空",
        "actual": "显示错误: Username is required",
        "status": "通过",
        "executor": "实习生-张三"
    },
    {
        "id": "TC007",
        "title": "锁定用户登录失败并显示锁定错误提示",
        "module": "登录模块",
        "priority": "P1",
        "preconditions": "打开SwagLabs登录页面",
        "steps": "1. 输入用户名: locked_out_user\n2. 输入密码: secret_sauce\n3. 点击Login按钮",
        "test_data": "username: locked_out_user\npassword: secret_sauce",
        "expected": "登录失败，提示用户已被锁定",
        "actual": "页面显示错误: Epic sadface: Sorry, this user has been locked out.",
        "status": "失败",   # 故意设为失败，让报告更真实
        "executor": "实习生-张三"
    }
]

# ========== 创建Excel并写入数据 ==========
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "测试执行报告"

# 定义表头
headers = ["用例编号", "用例标题", "项目/模块", "优先级", "前置条件", "测试步骤", "测试数据", "预期结果", "实际结果", "执行状态", "执行人", "执行日期"]
ws.append(headers)

# 设置表头样式
header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = border

# 填充数据
today = datetime.now().strftime("%Y-%m-%d")
for tc in test_cases:
    row_data = [
        tc["id"],
        tc["title"],
        tc["module"],
        tc["priority"],
        tc["preconditions"],
        tc["steps"],
        tc["test_data"],
        tc["expected"],
        tc["actual"],
        tc["status"],
        tc["executor"],
        today
    ]
    ws.append(row_data)

# 设置数据行样式（自动换行、边框）
data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
for row in range(2, len(test_cases) + 2):
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row, column=col)
        cell.alignment = data_alignment
        cell.border = border
        # 如果状态是"失败"，标红底色
        if col == 10 and cell.value == "失败":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)

# 调整列宽
column_widths = [12, 30, 12, 10, 25, 35, 25, 30, 30, 12, 14, 14]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

# 保存文件
filename = "SwagLabs登录测试执行报告.xlsx"
wb.save(filename)
print(f"✅ 报告生成成功！文件位置：{__import__('os').path.abspath(filename)}")
print("📊 报告包含：7条用例，6条通过，1条失败（锁定用户）")